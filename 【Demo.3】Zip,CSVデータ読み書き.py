
import zipfile

import pandas as pd
import time

import os
import io


#----- 入力ファイル -----
# Excel
READ_FILE_DIR = os.getcwd()
READ_FILE_PATH = READ_FILE_DIR + r'\csvのサンプル.zip'

#----- 出力ファイル -----
# Excel
WRITE_FILE_DIR = os.getcwd()
WRITE_FILE_PATH = READ_FILE_DIR + r'\csv_to_excel.xlsx'


#================================================================
# main()
#================================================================
def main():
    #------------------------------------------------
    # Zipファイル内のcsvファイルをDataFrameに読み込む
    #  [I] READ_FILE_PATH
    #  [O] list_csv
    #------------------------------------------------
    global list_csv
    list_csv = list()
    with zipfile.ZipFile(READ_FILE_PATH, 'r') as zf:
        for file in zf.namelist():
            df_csv_data = pd.read_csv(io.StringIO(zf.read(file).decode('cp932')))
            list_csv.append(df_csv_data) # リスト追加

    #------------------------------------------------
    # Excelファイル書き込み
    #  [I] df_output_data
    #  [O] WRITE_FILE_PATH
    #------------------------------------------------
    wb = None
    i = 0
    for df_csv in list_csv:
        wb = export_excel(_path=WRITE_FILE_PATH, _df=df_csv, _wb=wb, _col_width_=[12,10], _sheet_name='Sheet'+str(i))
        i += 1

    # ファイルを開く
    if True:
        import subprocess
        subprocess.Popen(['start', WRITE_FILE_PATH], shell=True)


##############################################################
# [export_excel] DataFrameをExcelへ出力(2021/01/31 12:26:01)     
#   画像を挿入するときは_wbはNoneでなければならない（_pathで指定する）
##############################################################
def export_excel(_path, _df=None, _wb=None, _sheet_name='sheet1', _letter_fmt=None, _append=False, _frz='B2', _auto_flt=True, _auto_sz=False, _header_height=None, _col_width_=[20,20], _header_fmt=None, _header_rot=0, _zoom=100, _heatmap=0, _is_index=True, _index_name='Index', _header_txtcol='000000', _header_fillcol='d9f2f8', _img=None):
    import openpyxl as px
    from openpyxl.utils import get_column_letter

    time_start = time.perf_counter()

    if _path==None and _wb==None:
        print('[Error] Both of _path and _bw are None.')
        return

    print('[Exporting Excel file ...] Sheet : "{0}"'.format(_sheet_name))
    
    #-------------------------------------------
    # 初期設定
    #-------------------------------------------
    # Workbook作成
    if _wb == None:
        if _append: # 既存ファイルにシート追加
            try:
                wb = px.load_workbook(_path)
            except:
                _append = False # ファイルが存在しないときは新規作成
        if not _append: # 新規ファイル
            wb = px.Workbook()
    else:
        wb = _wb
        _append = True
    # Worksheet作成
    ws = wb.create_sheet(title=_sheet_name)

    #-------------------------------------------
    # DataFrameをWorksheetに書き込み
    #-------------------------------------------
    if _df is not None:
        #----- 作業用にDataFrameをコピー -----
        df = _df.copy()
    
        # Timestampを文字列に変換（そのままだとエラーになるので）
        list_timestamp_col = list()
        # Timestampのセルが存在する列を探して文字列に変換する
        for col_name, col in df.iteritems():
            for item in col:
                tp = type(item)
                if tp is pd._libs.tslibs.timestamps.Timestamp:
                    list_timestamp_col.append(col_name)
                    break
        for col in list_timestamp_col:
            df[col] = df[col].astype(str)
            df[col] = df[col].replace('NaT', '')
    
        #----- Excelファイル用フォーマットの作成 -----
        base_font = '游ゴシック'
        from openpyxl.styles.fonts import Font
        from openpyxl.styles import PatternFill
        font_header_row = Font(name=base_font, b=True, sz=10, color=_header_txtcol)
        font_header_col = Font(name=base_font, b=True, sz=10, color=_header_txtcol)
        font_cell = Font(name=base_font, sz=10)
        align_header_row = px.styles.Alignment(horizontal="center", vertical="center", wrapText=True, textRotation=_header_rot)
        align_header_col = px.styles.Alignment(horizontal="center", vertical="center", wrapText=True)
        fill_header_row = PatternFill(patternType='solid', fgColor=_header_fillcol)
        fill_header_col = PatternFill(patternType='solid', fgColor=_header_fillcol)
    
        #----- データ出力 -----
        # DataFrameをWorksheetにExport
        l = df.columns.tolist()
        if _is_index:
            l.insert(0, _index_name) # 行のindexを先頭列に追加
        ws.append(l)
        count = 0
        for i, row in df.iterrows(): # 一行ずつwsに追加していく
            l = row.values.tolist()
            if _is_index:
                l.insert(0, row.name) # 行のindexを先頭列に追加
            ws.append(l)
            count += 1
            print('\r  - データコピー {0}/{1}'.format(count, len(df)), end="")
        print('')
    
        #-----  Worksheetの書式設定 -----
        # ヘッダー行(既定値)
        for cell in list(ws.rows)[0]:
            cell.font = font_header_row
            cell.alignment = align_header_row
            cell.fill = fill_header_row
        # ヘッダー行(個別)
        if _header_fmt != None:
            list_cell = list(ws.rows)[0]
            for head, fmt in _header_fmt.items():
                try:
                    index = list(df.columns).index(head)
                    if _is_index:
                        index += 1
                    cell = list_cell[index]
                except:
                    continue
                # rotation
                try:
                    rotation = fmt['rotation']
                    cell.alignment = px.styles.Alignment(horizontal="center", vertical="center", wrapText=True, textRotation=rotation)
                except:
                    pass
                # 文字色
                try:
                    text_color = fmt['txtcol']
                    cell.font = Font(name=base_font, b=True, sz=10, color=text_color)
                except:
                    pass
                # 背景色
                try:
                    fill_color = fmt['fillcol']
                    cell.fill = PatternFill(patternType='solid', fgColor=fill_color)
                except:
                    pass
        # 列ごとの書式設定用のリスト作成
        list_dtxt_pat = list()
        list_dfill_pat = list()
        if _header_fmt != None:
            for head, fmt in _header_fmt.items():
                try:
                    index = list(df.columns).index(head)
                    if _is_index:
                        index += 1
                except:
                    continue
                # 文字色
                try:
                    text_color = fmt['dtxtcol']
                    list_dtxt_pat.append([index, Font(name=base_font, sz=10, color=text_color)])
                except:
                    pass
                # 背景色
                try:
                    dfill_color = fmt['dfillcol']
                    list_dfill_pat.append([index, PatternFill(patternType='solid', fgColor=dfill_color)])
                except:
                    pass
        # データ行書式設定
        count = 0
        for row in ws.iter_rows(min_row=2): 
            # 書式設定
            for cell in row:
                cell.font = font_cell
            # 列ごとの書式設定で上書き
            for list_pat in list_dtxt_pat: # 個別設定がある列を順に処理する
                idx = list_pat[0]
                row[idx].font = list_pat[1]
            for list_pat in list_dfill_pat: # 個別設定がある列を順に処理する
                idx = list_pat[0]
                row[idx].fill = list_pat[1]
            # Index列がある場合はIndex用設定
            if _is_index:
                row[0].font = font_header_col # 先頭列のみ太字
                row[0].alignment = align_header_col # 先頭列のみセンタリング
                row[0].fill = fill_header_col # 先頭列の塗りつぶし
            count += 1
            print('\r  - 書式設定 {0}/{1}'.format(count, len(df)), end="")
        print('')
    
        #----- セルの文字書式 -----
        if type(_letter_fmt) is dict: # _header_fmtがあれば不要だが互換性のために残してある
            for col in ws.iter_cols():
                col_name = col[0].value
                if col_name in _letter_fmt:
                    num_format = _letter_fmt[col_name]
                    for cell in col:
                        cell.number_format = num_format
        elif type(_letter_fmt) is str:
            for col in ws.iter_cols():
                for cell in col:
                    cell.number_format = _letter_fmt
        # 列ごとの個別設定で上書き                
        if _header_fmt != None:
            list_col = list(_header_fmt.keys())
            for col in ws.iter_cols():
                col_name = col[0].value
                if col_name in list_col: # 列書式一覧の辞書にこの列が存在する
                    try:
                        fmt = _header_fmt[col_name]
                        num_format = fmt['dtxtformat']
                        for cell in col:
                            cell.number_format = num_format
                    except:
                        pass
       
        # Worksheetの列幅調整
        if _auto_sz: # 自動調整
            for col in ws.columns:
                max_length = 0
                column = col[0].column
                column = get_column_letter(column) # 数字をアルファベットに変換
                cols = col if _header_rot!=90 else col[1:]
                for cell in cols:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2) * 1.1
                ws.column_dimensions[column].width = adjusted_width
        else:
            for col in ws.columns:
                column = col[0].column # 列番号を取得
                col_letter = get_column_letter(column) # 列番号を列記号に変換
                width = _col_width_[0] if column == 1 else _col_width_[1] # 列幅
                ws.column_dimensions[col_letter].width = width
        # 列ごとの個別調整
        if _header_fmt != None:
            list_col = list(ws.columns)
            for head, fmt in _header_fmt.items():
                try:
                    width = fmt['width']
                    index = list(df.columns).index(head)
                    if _is_index:
                        index += 1
                    col = list_col[index]
                    column = col[0].column # 列番号を取得
                    col_letter = get_column_letter(column) # 列番号を列記号に変換
                    ws.column_dimensions[col_letter].width = width
                except:
                    pass
    
        # Worksheetの行の高さ調整
        if _header_height != None:
            ws.row_dimensions[1].height = _header_height
    
        # ヒートマップ
        from openpyxl.formatting.rule import ColorScale, FormatObject
        from openpyxl.styles import Color
        if _heatmap == 1: # 赤 → 白 → 青
            first = FormatObject(type='min')
            last = FormatObject(type='max')
            # colors match the format objects:
            colors = [Color('F8696B'), Color('5A8AC6')]
            # a three color scale would extend the sequences
            mid = FormatObject(type='percentile', val=50)
            colors.insert(1, Color('FCFCFF'))
            cs3 = ColorScale(cfvo=[first, mid, last], color=colors)
            # create a rule with the color scale
            from openpyxl.formatting.rule import Rule
            rule = Rule(type='colorScale', colorScale=cs3)
            # 対象範囲を示す文字列を作成
            rg = 'A2:' + get_column_letter(ws.max_column)+str(ws.max_row)
            ws.conditional_formatting.add(rg, rule)
        elif _heatmap == 2: # 白 → 橙 → 赤
            first = FormatObject(type='min')
            last = FormatObject(type='max')
            # colors match the format objects:
            colors = [Color('FFFFFF'), Color('F8696B')]
            # a three color scale would extend the sequences
            mid = FormatObject(type='percentile', val=50)
            colors.insert(1, Color('FFEB84'))
            cs3 = ColorScale(cfvo=[first, mid, last], color=colors)
            # create a rule with the color scale
            from openpyxl.formatting.rule import Rule
            rule = Rule(type='colorScale', colorScale=cs3)
            # 対象範囲を示す文字列を作成
            rg = 'A2:' + get_column_letter(ws.max_column)+str(ws.max_row)
            ws.conditional_formatting.add(rg, rule)
        elif _heatmap == 3: # 赤 → 橙 → 白
            first = FormatObject(type='min')
            last = FormatObject(type='max')
            # colors match the format objects:
            colors = [Color('F8696B'), Color('FFFFFF')]
            # a three color scale would extend the sequences
            mid = FormatObject(type='percentile', val=25)
            colors.insert(1, Color('FFEB84'))
            cs3 = ColorScale(cfvo=[first, mid, last], color=colors)
            # create a rule with the color scale
            from openpyxl.formatting.rule import Rule
            rule = Rule(type='colorScale', colorScale=cs3)
            # 対象範囲を示す文字列を作成
            rg = 'A2:' + get_column_letter(ws.max_column)+str(ws.max_row)
            ws.conditional_formatting.add(rg, rule)
            
        # 枠固定
        if _frz != None:
            ws.freeze_panes = _frz
        # オートフィルタ
        if _auto_flt:
            ws.auto_filter.ref = 'A1:' + get_column_letter(ws.max_column)+'1'
    
        # 表示倍率
        ws.sheet_view.zoomScale = _zoom
    
    #-------------------------------------------
    # Worksheetに画像を挿入
    #-------------------------------------------
    if _img != None:
        from openpyxl.drawing.image import Image
        for img in _img:
            fpath = img[0] # 挿入する画像ファイル
            anchor = img[1] # 挿入位置
            px_img = Image(fpath)
            px_img.anchor = anchor
            ws.add_image(px_img)
    
    #-------------------------------------------
    # Excelファイルに書き込み
    #-------------------------------------------
    # 最後に不要なシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    print('  - ファイル書き込み...', end='')
    wb.save(_path)
    # 画像ファイル削除
    if _img != None:
        for img in _img:
            is_delete = False # ファイルを削除するか否か
            if len(img) > 2:
                is_delete = img[2]
            if is_delete: # ファイル削除
                os.remove(fpath)

    print ('\n   ---> Finished. (処理時間:{0:.3f}[sec])'.format(time.perf_counter() - time_start ))
    
    return wb


##############################################################
# main proc
##############################################################
if __name__ == "__main__":
    main_time_start = time.perf_counter()
    main()
    print ("\n===> 正常終了 (処理時間:{0:.3f}[sec])".format(time.perf_counter() - main_time_start ))


